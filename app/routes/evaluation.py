from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, abort
from datetime import datetime, date, timedelta
from sqlalchemy import func, and_, or_, desc, asc
from app import db
from app.models import (Employee, Evaluation, TemplateEvaluation, CritereEvaluation, 
                       ObjectifEmploye, Utilisateur)
from app.forms import (EvaluationForm, TemplateEvaluationForm, CritereEvaluationForm, 
                      ObjectifEmployeForm, RapportEvaluationForm)
from app.utils.permissions import permission_requise
from flask_login import login_required, current_user
import json
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

evaluation_bp = Blueprint('evaluation', __name__)

# ============================================================================
# FONCTIONS UTILITAIRES
# ============================================================================

def _recalculer_score_global(evaluation):
    """Recalcule le score global d'une évaluation basé sur ses critères"""
    
    criteres = CritereEvaluation.query.filter_by(evaluation_id=evaluation.id).all()
    
    if not criteres:
        return
    
    # Calcul pondéré
    total_points = sum(c.score_obtenu * c.poids for c in criteres)
    total_max = sum(c.score_max * c.poids for c in criteres)
    
    if total_max > 0:
        score_global = (total_points / total_max) * 100
        evaluation.score_global = round(score_global, 2)
        
        # Mise à jour automatique de la note finale
        if score_global >= 90:
            evaluation.note_finale = 'Excellent'
        elif score_global >= 80:
            evaluation.note_finale = 'Très Bien'
        elif score_global >= 70:
            evaluation.note_finale = 'Bien'
        elif score_global >= 60:
            evaluation.note_finale = 'Satisfaisant'
        else:
            evaluation.note_finale = 'Insuffisant'

# ============================================================================
# DASHBOARD ET VUE PRINCIPALE
# ============================================================================

@evaluation_bp.route('/evaluations')
@login_required
@permission_requise('voir_evaluations')
def dashboard():
    """Dashboard principal du module Évaluations"""
    
    # Statistiques générales
    total_evaluations = Evaluation.query.count()
    evaluations_annee = Evaluation.query.filter(
        Evaluation.annee == datetime.now().year
    ).count()
    
    # Évaluations en attente (statut != Finalisé)
    evaluations_en_attente = Evaluation.query.filter(
        Evaluation.statut != 'Finalisé'
    ).count()
    
    # Moyenne des scores cette année
    avg_score = db.session.query(func.avg(Evaluation.score_global)).filter(
        Evaluation.annee == datetime.now().year
    ).scalar() or 0
    
    # Évaluations récentes
    evaluations_recentes = Evaluation.query.filter(
        Evaluation.date_creation >= datetime.now() - timedelta(days=30)
    ).order_by(desc(Evaluation.date_creation)).limit(5).all()
    
    # Employés à évaluer (pas d'évaluation cette année)
    employes_evalues = db.session.query(Evaluation.employe_id).filter(
        Evaluation.annee == datetime.now().year
    ).subquery()
    
    employes_a_evaluer = Employee.query.filter(
        Employee.statut == 'Actif',
        ~Employee.id.in_(employes_evalues)
    ).count()
    
    # Répartition par type d'évaluation
    repartition_types = db.session.query(
        Evaluation.type_evaluation,
        func.count(Evaluation.id).label('count')
    ).filter(
        Evaluation.annee == datetime.now().year
    ).group_by(Evaluation.type_evaluation).all()
    
    # Répartition par note finale
    repartition_notes = db.session.query(
        Evaluation.note_finale,
        func.count(Evaluation.id).label('count')
    ).filter(
        Evaluation.annee == datetime.now().year
    ).group_by(Evaluation.note_finale).all()
    
    return render_template('evaluations/dashboard.html',
                         total_evaluations=total_evaluations,
                         evaluations_annee=evaluations_annee,
                         evaluations_en_attente=evaluations_en_attente,
                         avg_score=round(avg_score, 2),
                         evaluations_recentes=evaluations_recentes,
                         employes_a_evaluer=employes_a_evaluer,
                         repartition_types=repartition_types,
                         repartition_notes=repartition_notes)

@evaluation_bp.route('/evaluations/list')
@login_required
@permission_requise('voir_evaluations')
def list_evaluations():
    """Liste toutes les évaluations avec filtres"""
    
    # Paramètres de filtrage
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    statut_filter = request.args.get('statut', '')
    type_filter = request.args.get('type', '')
    annee_filter = request.args.get('annee', '', type=str)
    departement_filter = request.args.get('departement', '')
    
    # Construction de la requête
    query = Evaluation.query.join(Employee)
    
    # Filtres de recherche
    if search:
        query = query.filter(
            or_(
                Employee.nom.ilike(f'%{search}%'),
                Employee.prenom.ilike(f'%{search}%'),
                Evaluation.periode.ilike(f'%{search}%')
            )
        )
    
    if statut_filter:
        query = query.filter(Evaluation.statut == statut_filter)
        
    if type_filter:
        query = query.filter(Evaluation.type_evaluation == type_filter)
        
    if annee_filter:
        try:
            annee = int(annee_filter)
            query = query.filter(Evaluation.annee == annee)
        except ValueError:
            pass
            
    if departement_filter:
        query = query.filter(Employee.departement == departement_filter)
    
    # Tri par date de création (plus récentes d'abord)
    query = query.order_by(desc(Evaluation.date_creation))
    
    # Pagination
    evaluations = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Données pour les filtres
    statuts = db.session.query(Evaluation.statut.distinct()).all()
    types = db.session.query(Evaluation.type_evaluation.distinct()).all()
    annees = db.session.query(Evaluation.annee.distinct()).order_by(desc(Evaluation.annee)).all()
    departements = db.session.query(Employee.departement.distinct()).filter(
        Employee.departement.isnot(None)
    ).all()
    
    return render_template('evaluations/list.html',
                         evaluations=evaluations,
                         search=search,
                         statut_filter=statut_filter,
                         type_filter=type_filter,
                         annee_filter=annee_filter,
                         departement_filter=departement_filter,
                         statuts=[s[0] for s in statuts if s[0]],
                         types=[t[0] for t in types if t[0]],
                         annees=[a[0] for a in annees if a[0]],
                         departements=[d[0] for d in departements if d[0]])

# ============================================================================
# GESTION DES ÉVALUATIONS
# ============================================================================

@evaluation_bp.route('/evaluations/new', methods=['GET', 'POST'])
@login_required
@permission_requise('creer_evaluations')
def create_evaluation():
    """Créer une nouvelle évaluation"""
    
    form = EvaluationForm()
    
    # Charger les listes déroulantes
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    form.employe_id.choices = [(e.id, f"{e.nom} {e.prenom or ''} - {e.poste or 'N/A'}") for e in employes]
    
    templates = TemplateEvaluation.query.filter_by(actif=True).order_by(TemplateEvaluation.nom).all()
    form.template_id.choices = [('', 'Aucun modèle')] + [(t.id, t.nom) for t in templates]
    
    if form.validate_on_submit():
        evaluation = Evaluation(
            employe_id=form.employe_id.data,
            evaluateur_id=current_user.id,
            template_id=form.template_id.data if form.template_id.data else None,
            periode=form.periode.data,
            type_evaluation=form.type_evaluation.data,
            annee=form.annee.data,
            score_global=form.score_global.data,
            note_finale=form.note_finale.data,
            objectifs_atteints=form.objectifs_atteints.data,
            objectifs_non_atteints=form.objectifs_non_atteints.data,
            objectifs_futurs=form.objectifs_futurs.data,
            plan_developpement=form.plan_developpement.data,
            formations_recommandees=form.formations_recommandees.data,
            commentaire_evaluateur=form.commentaire_evaluateur.data,
            commentaire_employe=form.commentaire_employe.data,
            commentaire_rh=form.commentaire_rh.data,
            date_evaluation=form.date_evaluation.data,
            statut=form.statut.data,
            created_by=current_user.id
        )
        
        db.session.add(evaluation)
        db.session.commit()
        
        flash('Évaluation créée avec succès', 'success')
        return redirect(url_for('evaluation.detail_evaluation', id=evaluation.id))
    
    return render_template('evaluations/form.html', form=form, title="Nouvelle Évaluation")

@evaluation_bp.route('/evaluations/<int:id>')
@login_required
@permission_requise('voir_evaluations')
def detail_evaluation(id):
    """Détail d'une évaluation"""
    
    evaluation = Evaluation.query.get_or_404(id)
    
    # Critères d'évaluation
    criteres = CritereEvaluation.query.filter_by(
        evaluation_id=id
    ).order_by(CritereEvaluation.section, CritereEvaluation.ordre).all()
    
    # Objectifs associés
    objectifs = ObjectifEmploye.query.filter_by(
        evaluation_id=id
    ).order_by(ObjectifEmploye.date_debut).all()
    
    # Grouper les critères par section
    criteres_par_section = {}
    for critere in criteres:
        if critere.section not in criteres_par_section:
            criteres_par_section[critere.section] = []
        criteres_par_section[critere.section].append(critere)
    
    return render_template('evaluations/detail.html',
                         evaluation=evaluation,
                         criteres_par_section=criteres_par_section,
                         objectifs=objectifs)

@evaluation_bp.route('/evaluations/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_requise('modifier_evaluations')
def edit_evaluation(id):
    """Modifier une évaluation"""
    
    evaluation = Evaluation.query.get_or_404(id)
    
    # Vérification des droits (seul l'évaluateur, l'admin ou RH peut modifier)
    if (evaluation.evaluateur_id != current_user.id and 
        not current_user.role.nom in ['Admin', 'Manager RH']):
        flash('Vous n\'avez pas les droits pour modifier cette évaluation', 'error')
        return redirect(url_for('evaluation.detail_evaluation', id=id))
    
    form = EvaluationForm(obj=evaluation)
    
    # Charger les listes déroulantes
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    form.employe_id.choices = [(e.id, f"{e.nom} {e.prenom or ''} - {e.poste or 'N/A'}") for e in employes]
    
    templates = TemplateEvaluation.query.filter_by(actif=True).order_by(TemplateEvaluation.nom).all()
    form.template_id.choices = [('', 'Aucun modèle')] + [(t.id, t.nom) for t in templates]
    
    if form.validate_on_submit():
        form.populate_obj(evaluation)
        evaluation.updated_by = current_user.id
        evaluation.updated_at = datetime.utcnow()
        
        # Mise à jour du statut de validation/finalisation
        if form.statut.data == 'Validé' and not evaluation.date_validation:
            evaluation.date_validation = datetime.utcnow()
        elif form.statut.data == 'Finalisé' and not evaluation.date_finalisation:
            evaluation.date_finalisation = datetime.utcnow()
        
        db.session.commit()
        
        flash('Évaluation modifiée avec succès', 'success')
        return redirect(url_for('evaluation.detail_evaluation', id=id))
    
    return render_template('evaluations/form.html', 
                         form=form, 
                         evaluation=evaluation,
                         title="Modifier l'Évaluation")

@evaluation_bp.route('/evaluations/<int:id>/delete', methods=['POST'])
@login_required
@permission_requise('supprimer_evaluations')
def delete_evaluation(id):
    """Supprimer une évaluation"""
    
    evaluation = Evaluation.query.get_or_404(id)
    
    # Vérification des droits
    if (evaluation.evaluateur_id != current_user.id and 
        not current_user.role.nom in ['Admin', 'Manager RH']):
        flash('Vous n\'avez pas les droits pour supprimer cette évaluation', 'error')
        return redirect(url_for('evaluation.detail_evaluation', id=id))
    
    # Supprimer les critères et objectifs associés
    CritereEvaluation.query.filter_by(evaluation_id=id).delete()
    ObjectifEmploye.query.filter_by(evaluation_id=id).update({'evaluation_id': None})
    
    employe_nom = f"{evaluation.employe.nom} {evaluation.employe.prenom or ''}"
    
    db.session.delete(evaluation)
    db.session.commit()
    
    flash(f'Évaluation de {employe_nom} supprimée avec succès', 'success')
    return redirect(url_for('evaluation.list_evaluations'))

# ============================================================================
# GESTION DES CRITÈRES D'ÉVALUATION
# ============================================================================

@evaluation_bp.route('/evaluations/<int:evaluation_id>/criteres')
@login_required
@permission_requise('voir_evaluations')
def list_criteres(evaluation_id):
    """Liste des critères d'une évaluation"""
    
    evaluation = Evaluation.query.get_or_404(evaluation_id)
    criteres = CritereEvaluation.query.filter_by(
        evaluation_id=evaluation_id
    ).order_by(CritereEvaluation.section, CritereEvaluation.ordre).all()
    
    # Grouper par section
    criteres_par_section = {}
    for critere in criteres:
        if critere.section not in criteres_par_section:
            criteres_par_section[critere.section] = []
        criteres_par_section[critere.section].append(critere)
    
    return render_template('evaluations/criteres.html',
                         evaluation=evaluation,
                         criteres_par_section=criteres_par_section)

@evaluation_bp.route('/evaluations/<int:evaluation_id>/criteres/add', methods=['GET', 'POST'])
@login_required
@permission_requise('modifier_evaluations')
def add_critere(evaluation_id):
    """Ajouter un critère à une évaluation"""
    
    evaluation = Evaluation.query.get_or_404(evaluation_id)
    form = CritereEvaluationForm()
    
    if form.validate_on_submit():
        critere = CritereEvaluation(
            evaluation_id=evaluation_id,
            section=form.section.data,
            critere=form.critere.data,
            description=form.description.data,
            score_obtenu=form.score_obtenu.data,
            score_max=form.score_max.data,
            poids=form.poids.data,
            commentaire=form.commentaire.data,
            recommandations=form.recommandations.data,
            ordre=form.ordre.data
        )
        
        db.session.add(critere)
        
        # Recalculer le score global
        _recalculer_score_global(evaluation)
        
        db.session.commit()
        
        flash('Critère ajouté avec succès', 'success')
        return redirect(url_for('evaluation.list_criteres', evaluation_id=evaluation_id))
    
    return render_template('evaluations/critere_form.html',
                         form=form,
                         evaluation=evaluation,
                         title="Ajouter un Critère")

@evaluation_bp.route('/evaluations/criteres/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_requise('modifier_evaluations')
def edit_critere(id):
    """Modifier un critère d'évaluation"""
    
    critere = CritereEvaluation.query.get_or_404(id)
    form = CritereEvaluationForm(obj=critere)
    
    if form.validate_on_submit():
        form.populate_obj(critere)
        
        # Recalculer le score global
        _recalculer_score_global(critere.evaluation)
        
        db.session.commit()
        
        flash('Critère modifié avec succès', 'success')
        return redirect(url_for('evaluation.list_criteres', evaluation_id=critere.evaluation_id))
    
    return render_template('evaluations/critere_form.html',
                         form=form,
                         critere=critere,
                         evaluation=critere.evaluation,
                         title="Modifier le Critère")

@evaluation_bp.route('/evaluations/criteres/<int:id>/delete', methods=['POST'])
@login_required
@permission_requise('modifier_evaluations')
def delete_critere(id):
    """Supprimer un critère d'évaluation"""
    
    critere = CritereEvaluation.query.get_or_404(id)
    evaluation_id = critere.evaluation_id
    evaluation = critere.evaluation
    
    db.session.delete(critere)
    
    # Recalculer le score global
    _recalculer_score_global(evaluation)
    
    db.session.commit()
    
    flash('Critère supprimé avec succès', 'success')
    return redirect(url_for('evaluation.list_criteres', evaluation_id=evaluation_id))

# ============================================================================
# GESTION DES OBJECTIFS
# ============================================================================

@evaluation_bp.route('/objectifs')
@login_required
@permission_requise('voir_evaluations')
def list_objectifs():
    """Liste tous les objectifs avec filtres"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    statut_filter = request.args.get('statut', '')
    employe_filter = request.args.get('employe', '', type=str)
    
    # Construction de la requête
    query = ObjectifEmploye.query.join(Employee)
    
    if search:
        query = query.filter(
            or_(
                ObjectifEmploye.titre.ilike(f'%{search}%'),
                ObjectifEmploye.description.ilike(f'%{search}%'),
                Employee.nom.ilike(f'%{search}%')
            )
        )
    
    if statut_filter:
        query = query.filter(ObjectifEmploye.statut == statut_filter)
        
    if employe_filter:
        try:
            employe_id = int(employe_filter)
            query = query.filter(ObjectifEmploye.employe_id == employe_id)
        except ValueError:
            pass
    
    query = query.order_by(desc(ObjectifEmploye.date_debut))
    
    objectifs = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Données pour les filtres
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    statuts = db.session.query(ObjectifEmploye.statut.distinct()).all()
    
    return render_template('evaluations/objectifs.html',
                         objectifs=objectifs,
                         employes=employes,
                         statuts=[s[0] for s in statuts if s[0]],
                         search=search,
                         statut_filter=statut_filter,
                         employe_filter=employe_filter)

@evaluation_bp.route('/objectifs/new', methods=['GET', 'POST'])
@login_required
@permission_requise('creer_evaluations')
def create_objectif():
    """Créer un nouvel objectif"""
    
    form = ObjectifEmployeForm()
    
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    form.employe_id.choices = [(e.id, f"{e.nom} {e.prenom or ''} - {e.poste or 'N/A'}") for e in employes]
    
    if form.validate_on_submit():
        objectif = ObjectifEmploye(
            employe_id=form.employe_id.data,
            titre=form.titre.data,
            description=form.description.data,
            type_objectif=form.type_objectif.data,
            priorite=form.priorite.data,
            date_debut=form.date_debut.data,
            date_fin=form.date_fin.data,
            date_realisation=form.date_realisation.data,
            indicateur_mesure=form.indicateur_mesure.data,
            valeur_cible=form.valeur_cible.data,
            valeur_atteinte=form.valeur_atteinte.data,
            pourcentage_realisation=form.pourcentage_realisation.data,
            statut=form.statut.data,
            resultat_obtenu=form.resultat_obtenu.data,
            commentaires=form.commentaires.data,
            created_by=current_user.id
        )
        
        db.session.add(objectif)
        db.session.commit()
        
        flash('Objectif créé avec succès', 'success')
        return redirect(url_for('evaluation.list_objectifs'))
    
    return render_template('evaluations/objectif_form.html', 
                         form=form, 
                         title="Nouvel Objectif")

@evaluation_bp.route('/objectifs/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_requise('modifier_evaluations')
def edit_objectif(id):
    """Modifier un objectif"""
    
    objectif = ObjectifEmploye.query.get_or_404(id)
    form = ObjectifEmployeForm(obj=objectif)
    
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    form.employe_id.choices = [(e.id, f"{e.nom} {e.prenom or ''} - {e.poste or 'N/A'}") for e in employes]
    
    if form.validate_on_submit():
        form.populate_obj(objectif)
        objectif.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('Objectif modifié avec succès', 'success')
        return redirect(url_for('evaluation.list_objectifs'))
    
    return render_template('evaluations/objectif_form.html',
                         form=form,
                         objectif=objectif,
                         title="Modifier l'Objectif")

@evaluation_bp.route('/objectifs/<int:id>/delete', methods=['POST'])
@login_required
@permission_requise('supprimer_evaluations')
def delete_objectif(id):
    """Supprimer un objectif"""
    
    objectif = ObjectifEmploye.query.get_or_404(id)
    
    db.session.delete(objectif)
    db.session.commit()
    
    flash('Objectif supprimé avec succès', 'success')
    return redirect(url_for('evaluation.list_objectifs'))

# ============================================================================
# GESTION DES TEMPLATES D'ÉVALUATION
# ============================================================================

@evaluation_bp.route('/templates')
@login_required
@permission_requise('administrer_evaluations')
def list_templates():
    """Liste des modèles d'évaluation"""
    
    templates = TemplateEvaluation.query.order_by(TemplateEvaluation.nom).all()
    
    return render_template('evaluations/templates.html', templates=templates)

@evaluation_bp.route('/templates/new', methods=['GET', 'POST'])
@login_required
@permission_requise('administrer_evaluations')
def create_template():
    """Créer un nouveau modèle d'évaluation"""
    
    form = TemplateEvaluationForm()
    
    if form.validate_on_submit():
        # Si marqué comme par défaut, désactiver les autres
        if form.par_defaut.data:
            TemplateEvaluation.query.filter_by(
                type_evaluation=form.type_evaluation.data,
                par_defaut=True
            ).update({'par_defaut': False})
        
        template = TemplateEvaluation(
            nom=form.nom.data,
            description=form.description.data,
            type_evaluation=form.type_evaluation.data,
            score_max=form.score_max.data,
            actif=form.actif.data,
            par_defaut=form.par_defaut.data,
            created_by=current_user.id
        )
        
        db.session.add(template)
        db.session.commit()
        
        flash('Modèle d\'évaluation créé avec succès', 'success')
        return redirect(url_for('evaluation.list_templates'))
    
    return render_template('evaluations/template_form.html',
                         form=form,
                         title="Nouveau Modèle")

@evaluation_bp.route('/templates/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_requise('administrer_evaluations')
def edit_template(id):
    """Modifier un modèle d'évaluation"""
    
    template = TemplateEvaluation.query.get_or_404(id)
    form = TemplateEvaluationForm(obj=template)
    
    if form.validate_on_submit():
        # Si marqué comme par défaut, désactiver les autres
        if form.par_defaut.data and not template.par_defaut:
            TemplateEvaluation.query.filter_by(
                type_evaluation=form.type_evaluation.data,
                par_defaut=True
            ).update({'par_defaut': False})
        
        form.populate_obj(template)
        template.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('Modèle d\'évaluation modifié avec succès', 'success')
        return redirect(url_for('evaluation.list_templates'))
    
    return render_template('evaluations/template_form.html',
                         form=form,
                         template=template,
                         title="Modifier le Modèle")

@evaluation_bp.route('/templates/<int:id>/delete', methods=['POST'])
@login_required
@permission_requise('administrer_evaluations')
def delete_template(id):
    """Supprimer un modèle d'évaluation"""
    
    template = TemplateEvaluation.query.get_or_404(id)
    
    # Vérifier s'il est utilisé
    evaluations_count = Evaluation.query.filter_by(template_id=id).count()
    
    if evaluations_count > 0:
        flash(f'Impossible de supprimer ce modèle : {evaluations_count} évaluation(s) l\'utilise(nt)', 'error')
        return redirect(url_for('evaluation.list_templates'))
    
    db.session.delete(template)
    db.session.commit()
    
    flash('Modèle d\'évaluation supprimé avec succès', 'success')
    return redirect(url_for('evaluation.list_templates'))

# ============================================================================
# RAPPORTS ET ANALYSES
# ============================================================================

@evaluation_bp.route('/rapports', methods=['GET', 'POST'])
@login_required
@permission_requise('voir_evaluations')
def rapports():
    """Interface de génération de rapports"""
    
    form = RapportEvaluationForm()
    
    # Charger les listes pour les filtres
    employes = Employee.query.filter_by(statut='Actif').order_by(Employee.nom).all()
    form.employes.choices = [(e.id, f"{e.nom} {e.prenom or ''}") for e in employes]
    
    departements = db.session.query(Employee.departement.distinct()).filter(
        Employee.departement.isnot(None)
    ).all()
    form.departements.choices = [(d[0], d[0]) for d in departements if d[0]]
    
    return render_template('evaluations/rapports.html', form=form)

@evaluation_bp.route('/rapports/generer', methods=['POST'])
@login_required
@permission_requise('voir_evaluations')
def generer_rapport():
    """Générer et télécharger un rapport d'évaluation"""
    
    form = RapportEvaluationForm()
    
    if not form.validate_on_submit():
        flash('Données du formulaire invalides', 'error')
        return redirect(url_for('evaluation.rapports'))
    
    # Construction de la requête
    query = Evaluation.query.join(Employee)
    
    # Filtres de date
    query = query.filter(
        Evaluation.date_evaluation >= form.date_debut.data,
        Evaluation.date_evaluation <= form.date_fin.data
    )
    
    # Filtres employés
    if form.employes.data:
        query = query.filter(Evaluation.employe_id.in_(form.employes.data))
    
    # Filtres types d'évaluation
    if form.types_evaluation.data:
        query = query.filter(Evaluation.type_evaluation.in_(form.types_evaluation.data))
    
    # Filtres départements
    if form.departements.data:
        query = query.filter(Employee.departement.in_(form.departements.data))
    
    evaluations = query.order_by(desc(Evaluation.date_evaluation)).all()
    
    # Génération selon le format
    if form.format_export.data == 'pdf':
        return _generer_rapport_pdf(evaluations, form)
    elif form.format_export.data == 'excel':
        return _generer_rapport_excel(evaluations, form)
    elif form.format_export.data == 'csv':
        return _generer_rapport_csv(evaluations, form)
    
    flash('Format de rapport non supporté', 'error')
    return redirect(url_for('evaluation.rapports'))

def _generer_rapport_pdf(evaluations, form):
    """Génère un rapport PDF"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.darkblue,
        alignment=1  # Centré
    )
    
    story.append(Paragraph("Rapport d'Évaluations", title_style))
    story.append(Spacer(1, 20))
    
    # Informations du rapport
    info_data = [
        ['Période:', f"{form.date_debut.data.strftime('%d/%m/%Y')} - {form.date_fin.data.strftime('%d/%m/%Y')}"],
        ['Type de rapport:', form.type_rapport.data],
        ['Nombre d\'évaluations:', str(len(evaluations))],
        ['Date de génération:', datetime.now().strftime('%d/%m/%Y à %H:%M')]
    ]
    
    info_table = Table(info_data, colWidths=[100, 300])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    if form.type_rapport.data == 'synthese':
        # Statistiques générales
        if evaluations:
            scores = [e.score_global for e in evaluations]
            avg_score = sum(scores) / len(scores)
            min_score = min(scores)
            max_score = max(scores)
            
            stats_data = [
                ['Statistiques', 'Valeur'],
                ['Score moyen:', f"{avg_score:.1f}"],
                ['Score minimum:', f"{min_score:.1f}"],
                ['Score maximum:', f"{max_score:.1f}"]
            ]
            
            stats_table = Table(stats_data, colWidths=[200, 100])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(Paragraph("Statistiques Générales", styles['Heading2']))
            story.append(stats_table)
            story.append(Spacer(1, 20))
    
    elif form.type_rapport.data == 'detaille':
        # Liste détaillée des évaluations
        story.append(Paragraph("Liste Détaillée des Évaluations", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        for evaluation in evaluations:
            eval_data = [
                ['Employé:', f"{evaluation.employe.nom} {evaluation.employe.prenom or ''}"],
                ['Période:', evaluation.periode],
                ['Type:', evaluation.type_evaluation],
                ['Score:', f"{evaluation.score_global}/100"],
                ['Note:', evaluation.note_finale],
                ['Date:', evaluation.date_evaluation.strftime('%d/%m/%Y')],
                ['Statut:', evaluation.statut]
            ]
            
            eval_table = Table(eval_data, colWidths=[80, 300])
            eval_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            
            story.append(eval_table)
            story.append(Spacer(1, 10))
    
    doc.build(story)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'rapport_evaluations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
        mimetype='application/pdf'
    )

def _generer_rapport_excel(evaluations, form):
    """Génère un rapport Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Évaluations"
    
    # En-têtes
    headers = [
        'ID', 'Employé', 'Poste', 'Département', 'Période', 'Type Évaluation',
        'Année', 'Score Global', 'Note Finale', 'Date Évaluation', 'Statut',
        'Évaluateur', 'Objectifs Atteints', 'Objectifs Non Atteints'
    ]
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    for row, evaluation in enumerate(evaluations, 2):
        ws.cell(row=row, column=1, value=evaluation.id)
        ws.cell(row=row, column=2, value=f"{evaluation.employe.nom} {evaluation.employe.prenom or ''}")
        ws.cell(row=row, column=3, value=evaluation.employe.poste or '')
        ws.cell(row=row, column=4, value=evaluation.employe.departement or '')
        ws.cell(row=row, column=5, value=evaluation.periode)
        ws.cell(row=row, column=6, value=evaluation.type_evaluation)
        ws.cell(row=row, column=7, value=evaluation.annee)
        ws.cell(row=row, column=8, value=evaluation.score_global)
        ws.cell(row=row, column=9, value=evaluation.note_finale)
        ws.cell(row=row, column=10, value=evaluation.date_evaluation.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=11, value=evaluation.statut)
        ws.cell(row=row, column=12, value=f"{evaluation.evaluateur.nom_complet or evaluation.evaluateur.nom_utilisateur}")
        ws.cell(row=row, column=13, value=evaluation.objectifs_atteints or '')
        ws.cell(row=row, column=14, value=evaluation.objectifs_non_atteints or '')
    
    # Ajustement automatique des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'rapport_evaluations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def _generer_rapport_csv(evaluations, form):
    """Génère un rapport CSV"""
    
    data = []
    
    # En-têtes
    headers = [
        'ID', 'Employé', 'Poste', 'Département', 'Période', 'Type Évaluation',
        'Année', 'Score Global', 'Note Finale', 'Date Évaluation', 'Statut'
    ]
    data.append(headers)
    
    # Données
    for evaluation in evaluations:
        row = [
            evaluation.id,
            f"{evaluation.employe.nom} {evaluation.employe.prenom or ''}",
            evaluation.employe.poste or '',
            evaluation.employe.departement or '',
            evaluation.periode,
            evaluation.type_evaluation,
            evaluation.annee,
            evaluation.score_global,
            evaluation.note_finale,
            evaluation.date_evaluation.strftime('%d/%m/%Y'),
            evaluation.statut
        ]
        data.append(row)
    
    # Création du DataFrame et export CSV
    df = pd.DataFrame(data[1:], columns=data[0])
    
    buffer = io.StringIO()
    df.to_csv(buffer, index=False, encoding='utf-8')
    buffer.seek(0)
    
    return send_file(
        io.BytesIO(buffer.getvalue().encode('utf-8')),
        as_attachment=True,
        download_name=f'rapport_evaluations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
        mimetype='text/csv'
    )

# ============================================================================
# API REST POUR INTÉGRATIONS
# ============================================================================

@evaluation_bp.route('/api/evaluations/stats')
@login_required
@permission_requise('voir_evaluations')
def api_evaluation_stats():
    """API pour les statistiques d'évaluations"""
    
    annee = request.args.get('annee', datetime.now().year, type=int)
    
    # Statistiques générales
    total = Evaluation.query.filter_by(annee=annee).count()
    avg_score = db.session.query(func.avg(Evaluation.score_global)).filter_by(annee=annee).scalar() or 0
    
    # Répartition par note
    repartition_notes = db.session.query(
        Evaluation.note_finale,
        func.count(Evaluation.id).label('count')
    ).filter_by(annee=annee).group_by(Evaluation.note_finale).all()
    
    # Évolution mensuelle
    evolution_mensuelle = db.session.query(
        func.extract('month', Evaluation.date_evaluation).label('mois'),
        func.count(Evaluation.id).label('count'),
        func.avg(Evaluation.score_global).label('avg_score')
    ).filter_by(annee=annee).group_by(
        func.extract('month', Evaluation.date_evaluation)
    ).order_by('mois').all()
    
    return jsonify({
        'total_evaluations': total,
        'score_moyen': round(avg_score, 2),
        'repartition_notes': [
            {'note': r[0], 'count': r[1]} for r in repartition_notes
        ],
        'evolution_mensuelle': [
            {
                'mois': int(r[0]),
                'nombre': r[1],
                'score_moyen': round(r[2] or 0, 2)
            } for r in evolution_mensuelle
        ]
    })

@evaluation_bp.route('/api/employes/<int:employe_id>/evaluations')
@login_required
@permission_requise('voir_evaluations')
def api_employe_evaluations(employe_id):
    """API pour récupérer les évaluations d'un employé"""
    
    employe = Employee.query.get_or_404(employe_id)
    
    evaluations = Evaluation.query.filter_by(
        employe_id=employe_id
    ).order_by(desc(Evaluation.date_evaluation)).all()
    
    data = []
    for evaluation in evaluations:
        data.append({
            'id': evaluation.id,
            'periode': evaluation.periode,
            'type_evaluation': evaluation.type_evaluation,
            'annee': evaluation.annee,
            'score_global': evaluation.score_global,
            'note_finale': evaluation.note_finale,
            'date_evaluation': evaluation.date_evaluation.isoformat(),
            'statut': evaluation.statut,
            'evaluateur': evaluation.evaluateur.nom_complet or evaluation.evaluateur.nom_utilisateur
        })
    
    return jsonify({
        'employe': {
            'id': employe.id,
            'nom': employe.nom,
            'prenom': employe.prenom,
            'poste': employe.poste
        },
        'evaluations': data,
        'total': len(data)
    })
